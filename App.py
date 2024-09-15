import pyautogui
import webbrowser
import openpyxl
from urllib.parse import quote   # Biblioteca que permite a formatação de links especiais
from time import sleep


""""
Automatizar o Envio de Mensagens para o pagamento de boletos.
"""

webbrowser.open('https://web.whatsapp.com/')
sleep(30)

# Descrição das Etapas.

# Ler a planilha e guardar informações sobre nome, telefone e data de vencimento.

workbook = openpyxl.load_workbook('Vendas.xlsx')        # Abre o arquivo do Excel

pagina_clientes = workbook['Planilha1']                 # Seleciona a aba da planilha que contem as informações

for linha in pagina_clientes.iter_rows(min_row=2):      # iter_rows (Função especifica para passsar pelas linhas e ler as informações, começando pela linha 2)
    
    # Extrair nome, telefone e vencimento
    
    nome = linha[0].value
    telefone = linha[1].value
    vencimento = linha[2].value
    
    link = 'https://testedeautomacao.com.br'
    mensagem = f'olá, {nome}. Seu boleto vence no dia {vencimento.strftime('%d/%m/%Y')}. Pague pelo link {link}'
    
    # Tratamento de erros em que caso a mensagem não seja enviada, um arquivo 'erros.csv' será gerado, 
    # contendo o nome e telefone dos clientes que não receberam a mensagem
    
    try:
        # Criar links personalizado do whatsapp 
        
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
        
        # Enviar mensagens para cada cliente com base nos dados da planilha.
        
        webbrowser.open(link_mensagem_whatsapp)
        sleep(10)
        seta = pyautogui.locateCenterOnScreen('Seta.png')
        sleep(5)
        pyautogui.click(seta[0],seta[1])
        sleep(5)
        pyautogui.hotkey('ctrl','w')
        sleep(5)
    except:
        print(f'Não foi possivel enviar a mensagem para {nome}')
        with open('erros.csv', 'a',newline='', encoding='utf-8') as arquivo:
            arquivo.write(f'{nome},{telefone}')

